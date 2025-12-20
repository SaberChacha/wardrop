"""
Generate dummy Excel data for Wardrop Dashboard
Run: python generate_dummy_data.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random
from datetime import datetime, timedelta

# Styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="B76E79", end_color="B76E79", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws):
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def auto_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================
# CLIENTS DATA
# ============================================
def generate_clients():
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    # Headers (matching import format)
    headers = ["ID", "Full Name", "Phone", "WhatsApp", "Address", "Notes"]
    ws.append(headers)
    style_header(ws)
    
    # Algerian names
    first_names_f = ["Amina", "Fatima", "Khadija", "Meriem", "Sara", "Yasmine", "Nadia", "Leila", "Samira", "Houria", 
                    "Nour", "Rania", "Imane", "Soumia", "Lamia", "Asma", "Hanane", "Djamila", "Karima", "Malika"]
    last_names = ["Benali", "Belkacem", "Bouzid", "Khelifa", "Amrani", "Hadj", "Messaoud", "Saidi", "Mebarki", "Ferhat",
                  "Boudiaf", "Aissaoui", "Hamdi", "Taleb", "Zerrouki", "Rahmani", "Slimani", "Djamel", "Mokrani", "Chaouch"]
    
    wilayas = ["Alger", "Oran", "Constantine", "Annaba", "Blida", "S√©tif", "Tlemcen", "B√©ja√Øa", "Tizi Ouzou", "Batna"]
    
    clients_data = []
    for i in range(30):
        name = f"{random.choice(first_names_f)} {random.choice(last_names)}"
        phone = f"0{random.choice(['5', '6', '7'])}{random.randint(10000000, 99999999)}"
        whatsapp = phone
        wilaya = random.choice(wilayas)
        address = f"{random.randint(1, 200)} Rue {random.choice(['des Martyrs', 'Didouche Mourad', 'Ben Mhidi', 'Emir Abdelkader', 'du 1er Novembre'])}, {wilaya}"
        notes_options = ["", "Cliente fid√®le", "Recommand√©e par amie", "Mariage pr√©vu", "Premier contact", "Cliente VIP", ""]
        notes = random.choice(notes_options)
        
        clients_data.append([i+1, name, phone, whatsapp, address, notes])
    
    for row in clients_data:
        ws.append(row)
    
    auto_width(ws)
    wb.save("clients_import.xlsx")
    print("‚úì Generated: clients_import.xlsx (30 clients)")

# ============================================
# DRESSES DATA
# ============================================
def generate_dresses():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dresses"
    
    # Headers (matching import format)
    headers = ["ID", "Name", "Category", "Size", "Color", "Rental Price (DZD)", "Deposit (DZD)", "Status", "Description"]
    ws.append(headers)
    style_header(ws)
    
    categories = ["Robe de Mari√©e", "Robe de Soir√©e", "Robe de Fian√ßailles", "Caftan", "Karakou", "Robe Kabyle"]
    sizes = ["XS", "S", "M", "L", "XL", "XXL"]
    colors = ["Blanc", "Ivoire", "Champagne", "Rose Poudr√©", "Or", "Argent√©", "Rouge Bordeaux", "Bleu Royal", "Vert √âmeraude", "Noir"]
    statuses = ["available", "available", "available", "rented", "available", "maintenance"]
    
    dress_names = [
        "Princesse des Neiges", "√âtoile du Soir", "Rose de Kabylie", "Sultane d'Or", "Perle de l'Atlas",
        "Lune de Miel", "Fleur de Jasmin", "Reine du D√©sert", "√âclat de Diamant", "Belle du Sahel",
        "Aurore Dor√©e", "Cam√©lia Blanc", "Orchid√©e Royale", "Gazelle du Maghreb", "Nuit √âtoil√©e",
        "Dune d'Or", "Palais des R√™ves", "Cascade de Soie", "Mirage Enchant√©", "Tr√©sor Berb√®re"
    ]
    
    dresses_data = []
    for i, name in enumerate(dress_names):
        category = random.choice(categories)
        size = random.choice(sizes)
        color = random.choice(colors)
        
        # Price based on category
        if category in ["Robe de Mari√©e", "Karakou"]:
            rental_price = random.randint(25000, 80000)
            deposit = random.randint(10000, 30000)
        elif category == "Caftan":
            rental_price = random.randint(15000, 45000)
            deposit = random.randint(8000, 20000)
        else:
            rental_price = random.randint(10000, 35000)
            deposit = random.randint(5000, 15000)
        
        status = random.choice(statuses)
        descriptions = [
            f"Magnifique {category.lower()} brod√©e √† la main avec des perles",
            f"√âl√©gante {category.lower()} avec dentelle import√©e",
            f"{category} traditionnelle avec finitions dor√©es",
            f"Cr√©ation exclusive, taille ajustable",
            f"Design moderne inspir√© de la tradition alg√©rienne"
        ]
        description = random.choice(descriptions)
        
        dresses_data.append([i+1, name, category, size, color, rental_price, deposit, status, description])
    
    for row in dresses_data:
        ws.append(row)
    
    auto_width(ws)
    wb.save("dresses_import.xlsx")
    print("‚úì Generated: dresses_import.xlsx (20 dresses)")

# ============================================
# CLOTHING DATA
# ============================================
def generate_clothing():
    wb = Workbook()
    ws = wb.active
    ws.title = "Clothing"
    
    # Headers (matching import format)
    headers = ["ID", "Name", "Category", "Size", "Color", "Sale Price (DZD)", "Stock Quantity", "Description"]
    ws.append(headers)
    style_header(ws)
    
    categories = ["Hijab", "Abaya", "Accessoires", "Ch√¢le", "Foulard", "Bijoux", "Ceinture", "Voile"]
    sizes = ["Unique", "S", "M", "L", "XL"]
    colors = ["Blanc", "Noir", "Beige", "Rose", "Bleu Marine", "Gris", "Bordeaux", "Vert Sapin", "Dor√©", "Argent√©"]
    
    clothing_items = [
        ("Hijab Satin Premium", "Hijab", 2500, 3500),
        ("Hijab Jersey Confort", "Hijab", 1500, 2500),
        ("Abaya Brod√©e Dubai", "Abaya", 8000, 15000),
        ("Abaya Simple √âl√©gante", "Abaya", 5000, 9000),
        ("Ch√¢le Pashmina", "Ch√¢le", 3000, 5000),
        ("Foulard Soie Naturelle", "Foulard", 4000, 7000),
        ("Parure Bijoux Mariage", "Bijoux", 6000, 12000),
        ("Boucles Perles", "Bijoux", 2000, 4000),
        ("Ceinture Dor√©e", "Ceinture", 3500, 6000),
        ("Voile de Mari√©e", "Voile", 5000, 10000),
        ("Hijab Mousseline", "Hijab", 1800, 3000),
        ("Abaya Papillon", "Abaya", 7000, 11000),
        ("Bandeau Strass", "Accessoires", 1500, 3000),
        ("Ch√¢le Brod√© Main", "Ch√¢le", 4500, 8000),
        ("Collier Tradition", "Bijoux", 5500, 9000),
    ]
    
    clothing_data = []
    for i, (name, category, min_price, max_price) in enumerate(clothing_items):
        size = "Unique" if category in ["Hijab", "Foulard", "Bijoux", "Accessoires", "Ch√¢le", "Voile"] else random.choice(sizes)
        color = random.choice(colors)
        price = random.randint(min_price, max_price)
        stock = random.randint(2, 15)
        descriptions = [
            "Qualit√© sup√©rieure, import√©",
            "Fait main, pi√®ce unique",
            "Collection 2025",
            "Best-seller",
            "Nouveau arrivage"
        ]
        description = random.choice(descriptions)
        
        clothing_data.append([i+1, name, category, size, color, price, stock, description])
    
    for row in clothing_data:
        ws.append(row)
    
    auto_width(ws)
    wb.save("clothing_import.xlsx")
    print("‚úì Generated: clothing_import.xlsx (15 items)")

# ============================================
# RUN ALL
# ============================================
if __name__ == "__main__":
    print("\nüéÄ Generating Wardrop Sample Data...\n")
    generate_clients()
    generate_dresses()
    generate_clothing()
    print("\n‚úÖ All files generated successfully!")
    print("\nYou can import these files through the Wardrop dashboard:")
    print("  ‚Ä¢ Settings ‚Üí Import ‚Üí Select file type ‚Üí Upload")
    print("\nOr via API:")
    print("  ‚Ä¢ POST /api/export/import/clients")
    print("  ‚Ä¢ POST /api/export/import/dresses")
    print("  ‚Ä¢ POST /api/export/import/clothing")



