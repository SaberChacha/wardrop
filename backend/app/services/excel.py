from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date, datetime
from io import BytesIO

from ..models.client import Client
from ..models.dress import Dress
from ..models.clothing import Clothing
from ..models.booking import Booking
from ..models.sale import Sale
from ..models.expense import Expense

# Translation dictionaries for Excel exports
TRANSLATIONS = {
    "ar": {
        "summary": "ملخص",
        "commercial_report": "التقرير التجاري",
        "period": "الفترة",
        "metric": "المقياس",
        "value": "القيمة",
        "total_rental_revenue": "إجمالي إيرادات التأجير (د.ج)",
        "total_sales_revenue": "إجمالي إيرادات المبيعات (د.ج)",
        "total_revenue": "إجمالي الإيرادات (د.ج)",
        "total_expenses": "إجمالي المصاريف (د.ج)",
        "net_profit": "صافي الربح (د.ج)",
        "number_of_bookings": "عدد الحجوزات",
        "number_of_sales": "عدد المبيعات",
        "number_of_expenses": "عدد المصاريف",
        "bookings": "الحجوزات",
        "sales": "المبيعات",
        "expenses": "المصاريف",
        "client": "العميل",
        "dress": "الفستان",
        "item": "المنتج",
        "start_date": "تاريخ البداية",
        "end_date": "تاريخ النهاية",
        "price": "السعر (د.ج)",
        "deposit": "الضمان (د.ج)",
        "status": "الحالة",
        "quantity": "الكمية",
        "unit_price": "سعر الوحدة (د.ج)",
        "total": "الإجمالي (د.ج)",
        "date": "التاريخ",
        "reason": "السبب",
        "amount": "المبلغ (د.ج)",
    },
    "fr": {
        "summary": "Résumé",
        "commercial_report": "Rapport Commercial",
        "period": "Période",
        "metric": "Métrique",
        "value": "Valeur",
        "total_rental_revenue": "Total Revenus Location (DZD)",
        "total_sales_revenue": "Total Revenus Ventes (DZD)",
        "total_revenue": "Total Revenus (DZD)",
        "total_expenses": "Total Dépenses (DZD)",
        "net_profit": "Bénéfice Net (DZD)",
        "number_of_bookings": "Nombre de Réservations",
        "number_of_sales": "Nombre de Ventes",
        "number_of_expenses": "Nombre de Dépenses",
        "bookings": "Réservations",
        "sales": "Ventes",
        "expenses": "Dépenses",
        "client": "Client",
        "dress": "Robe",
        "item": "Article",
        "start_date": "Date de Début",
        "end_date": "Date de Fin",
        "price": "Prix (DZD)",
        "deposit": "Dépôt (DZD)",
        "status": "Statut",
        "quantity": "Quantité",
        "unit_price": "Prix Unitaire (DZD)",
        "total": "Total (DZD)",
        "date": "Date",
        "reason": "Motif",
        "amount": "Montant (DZD)",
    }
}


class ExcelService:
    def __init__(self, db: Session):
        self.db = db
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="B76E79", end_color="B76E79", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_header(self, ws, row=1):
        for cell in ws[row]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

    def _auto_width(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_clients(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"

        # Headers
        headers = ["ID", "Full Name", "Phone", "WhatsApp", "Address", "Notes", "Created At"]
        ws.append(headers)
        self._style_header(ws)

        # Data
        clients = self.db.query(Client).order_by(Client.full_name).all()
        for client in clients:
            ws.append([
                client.id,
                client.full_name,
                client.phone,
                client.whatsapp,
                client.address,
                client.notes,
                client.created_at.strftime("%Y-%m-%d %H:%M") if client.created_at else ""
            ])

        self._auto_width(ws)
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_dresses(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dresses"

        headers = ["ID", "Name", "Category", "Size", "Color", "Rental Price (DZD)", "Deposit (DZD)", "Status", "Description", "Created At"]
        ws.append(headers)
        self._style_header(ws)

        dresses = self.db.query(Dress).order_by(Dress.name).all()
        for dress in dresses:
            ws.append([
                dress.id,
                dress.name,
                dress.category,
                dress.size,
                dress.color,
                float(dress.rental_price),
                float(dress.deposit_amount),
                dress.status,
                dress.description,
                dress.created_at.strftime("%Y-%m-%d %H:%M") if dress.created_at else ""
            ])

        self._auto_width(ws)
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_clothing(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clothing"

        headers = ["ID", "Name", "Category", "Size", "Color", "Sale Price (DZD)", "Stock Quantity", "Description", "Created At"]
        ws.append(headers)
        self._style_header(ws)

        items = self.db.query(Clothing).order_by(Clothing.name).all()
        for item in items:
            ws.append([
                item.id,
                item.name,
                item.category,
                item.size,
                item.color,
                float(item.sale_price),
                item.stock_quantity,
                item.description,
                item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else ""
            ])

        self._auto_width(ws)
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_bookings(self, start_date: Optional[date] = None, end_date: Optional[date] = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"

        headers = ["ID", "Client", "Dress", "Start Date", "End Date", "Rental Price (DZD)", "Deposit (DZD)", "Deposit Status", "Booking Status", "Notes", "Created At"]
        ws.append(headers)
        self._style_header(ws)

        query = self.db.query(Booking)
        if start_date:
            query = query.filter(Booking.start_date >= start_date)
        if end_date:
            query = query.filter(Booking.end_date <= end_date)
        
        bookings = query.order_by(Booking.start_date.desc()).all()
        for booking in bookings:
            ws.append([
                booking.id,
                booking.client.full_name if booking.client else "",
                booking.dress.name if booking.dress else "",
                booking.start_date.strftime("%Y-%m-%d") if booking.start_date else "",
                booking.end_date.strftime("%Y-%m-%d") if booking.end_date else "",
                float(booking.rental_price),
                float(booking.deposit_amount),
                booking.deposit_status,
                booking.booking_status,
                booking.notes,
                booking.created_at.strftime("%Y-%m-%d %H:%M") if booking.created_at else ""
            ])

        self._auto_width(ws)
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_sales(self, start_date: Optional[date] = None, end_date: Optional[date] = None) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        headers = ["ID", "Client", "Item", "Quantity", "Unit Price (DZD)", "Total Price (DZD)", "Sale Date", "Notes", "Created At"]
        ws.append(headers)
        self._style_header(ws)

        query = self.db.query(Sale)
        if start_date:
            query = query.filter(Sale.sale_date >= start_date)
        if end_date:
            query = query.filter(Sale.sale_date <= end_date)
        
        sales = query.order_by(Sale.sale_date.desc()).all()
        for sale in sales:
            ws.append([
                sale.id,
                sale.client.full_name if sale.client else "",
                sale.clothing.name if sale.clothing else "",
                sale.quantity,
                float(sale.unit_price),
                float(sale.total_price),
                sale.sale_date.strftime("%Y-%m-%d") if sale.sale_date else "",
                sale.notes,
                sale.created_at.strftime("%Y-%m-%d %H:%M") if sale.created_at else ""
            ])

        self._auto_width(ws)
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_commercial_report(self, start_date: Optional[date] = None, end_date: Optional[date] = None, lang: str = "fr") -> bytes:
        wb = Workbook()
        
        # Get translations for the selected language
        t = TRANSLATIONS.get(lang, TRANSLATIONS["fr"])
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = t["summary"]
        
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = date(end_date.year, 1, 1)  # Start of year
        
        # Calculate totals
        from sqlalchemy import func
        
        rental_total = self.db.query(func.sum(Booking.rental_price)).filter(
            Booking.start_date >= start_date,
            Booking.start_date <= end_date,
            Booking.booking_status != "cancelled"
        ).scalar() or 0
        
        sales_total = self.db.query(func.sum(Sale.total_price)).filter(
            Sale.sale_date >= start_date,
            Sale.sale_date <= end_date
        ).scalar() or 0
        
        expenses_total = self.db.query(func.sum(Expense.amount)).filter(
            Expense.date >= start_date,
            Expense.date <= end_date
        ).scalar() or 0
        
        booking_count = self.db.query(Booking).filter(
            Booking.start_date >= start_date,
            Booking.start_date <= end_date,
            Booking.booking_status != "cancelled"
        ).count()
        
        sales_count = self.db.query(Sale).filter(
            Sale.sale_date >= start_date,
            Sale.sale_date <= end_date
        ).count()
        
        expenses_count = self.db.query(Expense).filter(
            Expense.date >= start_date,
            Expense.date <= end_date
        ).count()
        
        total_revenue = float(rental_total + sales_total)
        net_profit = total_revenue - float(expenses_total)
        
        ws_summary.append([t["commercial_report"]])
        ws_summary.append([f"{t['period']}: {start_date} - {end_date}"])
        ws_summary.append([])
        ws_summary.append([t["metric"], t["value"]])
        self._style_header(ws_summary, 4)
        ws_summary.append([t["total_rental_revenue"], float(rental_total)])
        ws_summary.append([t["total_sales_revenue"], float(sales_total)])
        ws_summary.append([t["total_revenue"], total_revenue])
        ws_summary.append([t["total_expenses"], float(expenses_total)])
        ws_summary.append([t["net_profit"], net_profit])
        ws_summary.append([t["number_of_bookings"], booking_count])
        ws_summary.append([t["number_of_sales"], sales_count])
        ws_summary.append([t["number_of_expenses"], expenses_count])
        
        self._auto_width(ws_summary)
        
        # Bookings Sheet
        ws_bookings = wb.create_sheet(t["bookings"])
        headers = [t["client"], t["dress"], t["start_date"], t["end_date"], t["price"], t["deposit"], t["status"]]
        ws_bookings.append(headers)
        self._style_header(ws_bookings)
        
        bookings = self.db.query(Booking).filter(
            Booking.start_date >= start_date,
            Booking.start_date <= end_date
        ).order_by(Booking.start_date.desc()).all()
        
        for booking in bookings:
            ws_bookings.append([
                booking.client.full_name if booking.client else "",
                booking.dress.name if booking.dress else "",
                booking.start_date.strftime("%Y-%m-%d"),
                booking.end_date.strftime("%Y-%m-%d"),
                float(booking.rental_price),
                float(booking.deposit_amount),
                booking.booking_status
            ])
        
        self._auto_width(ws_bookings)
        
        # Sales Sheet
        ws_sales = wb.create_sheet(t["sales"])
        headers = [t["client"], t["item"], t["quantity"], t["unit_price"], t["total"], t["date"]]
        ws_sales.append(headers)
        self._style_header(ws_sales)
        
        sales = self.db.query(Sale).filter(
            Sale.sale_date >= start_date,
            Sale.sale_date <= end_date
        ).order_by(Sale.sale_date.desc()).all()
        
        for sale in sales:
            ws_sales.append([
                sale.client.full_name if sale.client else "",
                sale.clothing.name if sale.clothing else "",
                sale.quantity,
                float(sale.unit_price),
                float(sale.total_price),
                sale.sale_date.strftime("%Y-%m-%d")
            ])
        
        self._auto_width(ws_sales)
        
        # Expenses Sheet
        ws_expenses = wb.create_sheet(t["expenses"])
        headers = [t["date"], t["reason"], t["amount"]]
        ws_expenses.append(headers)
        self._style_header(ws_expenses)
        
        expenses = self.db.query(Expense).filter(
            Expense.date >= start_date,
            Expense.date <= end_date
        ).order_by(Expense.date.desc()).all()
        
        for expense in expenses:
            ws_expenses.append([
                expense.date.strftime("%Y-%m-%d"),
                expense.reason,
                float(expense.amount)
            ])
        
        self._auto_width(ws_expenses)
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def import_clients(self, file_contents: bytes) -> dict:
        wb = load_workbook(filename=BytesIO(file_contents))
        ws = wb.active
        
        imported = 0
        errors = []
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx, row in enumerate(rows, start=2):
            try:
                if not row[1]:  # Skip if no name
                    continue
                    
                client = Client(
                    full_name=str(row[1]) if row[1] else "",
                    phone=str(row[2]) if row[2] else None,
                    whatsapp=str(row[3]) if row[3] else None,
                    address=str(row[4]) if row[4] else None,
                    notes=str(row[5]) if row[5] else None
                )
                self.db.add(client)
                imported += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        self.db.commit()
        return {"imported": imported, "errors": errors}

    def import_dresses(self, file_contents: bytes) -> dict:
        wb = load_workbook(filename=BytesIO(file_contents))
        ws = wb.active
        
        imported = 0
        errors = []
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx, row in enumerate(rows, start=2):
            try:
                if not row[1]:  # Skip if no name
                    continue
                    
                dress = Dress(
                    name=str(row[1]) if row[1] else "",
                    category=str(row[2]) if row[2] else "Other",
                    size=str(row[3]) if row[3] else "M",
                    color=str(row[4]) if row[4] else "White",
                    rental_price=float(row[5]) if row[5] else 0,
                    deposit_amount=float(row[6]) if row[6] else 0,
                    status=str(row[7]) if row[7] else "available",
                    description=str(row[8]) if row[8] else None
                )
                self.db.add(dress)
                imported += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        self.db.commit()
        return {"imported": imported, "errors": errors}

    def import_clothing(self, file_contents: bytes) -> dict:
        wb = load_workbook(filename=BytesIO(file_contents))
        ws = wb.active
        
        imported = 0
        errors = []
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx, row in enumerate(rows, start=2):
            try:
                if not row[1]:  # Skip if no name
                    continue
                    
                item = Clothing(
                    name=str(row[1]) if row[1] else "",
                    category=str(row[2]) if row[2] else "Other",
                    size=str(row[3]) if row[3] else "M",
                    color=str(row[4]) if row[4] else "Black",
                    sale_price=float(row[5]) if row[5] else 0,
                    stock_quantity=int(row[6]) if row[6] else 0,
                    description=str(row[7]) if row[7] else None
                )
                self.db.add(item)
                imported += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        self.db.commit()
        return {"imported": imported, "errors": errors}

